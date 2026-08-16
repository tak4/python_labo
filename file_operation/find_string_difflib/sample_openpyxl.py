import openpyxl

wb = openpyxl.load_workbook("excel_workbooks/sample001.xlsx") # Excelファイルの読み込み

sheet_names = wb.sheetnames # シート名一覧取得
print(sheet_names)

ws = wb[sheet_names[0]] # シートの取得
cell = ws.cell(row=1, column=1) # A1セルの取得
print(cell.value)


result = []
for row in range(1, ws.max_row + 1):
    value = ws.cell(row=row, column=1).value

    # None か 空文字なら除外
    if value is not None and str(value).strip() != "":
        result.append((row, value))

print(result)

result = []
for row in ws: # 各行でループ
    for cell in row:
        result.append(cell.value)

print(result)
